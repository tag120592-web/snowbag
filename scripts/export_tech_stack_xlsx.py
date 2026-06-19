#!/usr/bin/env python3
"""Export Snowbag technology stack to Excel."""
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path.home() / "Desktop" / "snowbag_tech_stack.xlsx"

header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="2F5496")
section_font = Font(bold=True, size=12, color="2F5496")
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")


def style_header_row(ws, row: int, cols: int) -> None:
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border


def write_table(ws, start_row: int, headers: list[str], rows: list[tuple], col_widths: list[int]) -> int:
    for i, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=i, value=h)
    style_header_row(ws, start_row, len(headers))
    r = start_row + 1
    for row_data in rows:
        for i, val in enumerate(row_data, 1):
            cell = ws.cell(row=r, column=i, value=val)
            cell.alignment = wrap
            cell.border = border
        r += 1
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    return r


def main() -> None:
    wb = Workbook()

    ws = wb.active
    ws.title = "Сводка"
    ws["A1"] = "Стек технологий — Формула расчета"
    ws["A1"].font = Font(bold=True, size=16, color="2F5496")
    ws["A2"] = f"Дата: {date.today().strftime('%d.%m.%Y')}"
    ws["A3"] = (
        "Vue 3 + TypeScript + Vite + Three.js → Go (chi) REST API → "
        "PostgreSQL 16 + PostGIS, Redis, MinIO, Docker Compose"
    )
    ws["A3"].alignment = wrap
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")
    ws.merge_cells("A3:D3")
    ws.row_dimensions[3].height = 36

    r = 5
    ws.cell(row=r, column=1, value="Слой").font = section_font
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
    r += 1
    summary = [
        ("Frontend", "Vue 3, TypeScript, Vite 6, Vue Router 4, Three.js, pdfjs-dist, Canvas/SVG, Яндекс.Карты"),
        ("Backend", "Go 1.22, chi v5, pgx v5, minio-go, gofpdf, excelize"),
        ("БД", "PostgreSQL 16, PostGIS 3.4, JSONB, pgcrypto"),
        ("Инфраструктура", "Docker Compose, Redis 7, MinIO, nginx 1.27, Alpine Linux"),
        ("Интеграции", "Яндекс.Карты, Nominatim, ПИМ QRC API (ТехноНИКОЛЬ)"),
        ("Нормативы", "СП 20.13330.2016, СНиП (климат), СП 230 (теплотехника)"),
    ]
    write_table(ws, r, ["Компонент", "Технологии"], summary, [22, 70])

    ws2 = wb.create_sheet("Frontend")
    write_table(
        ws2,
        1,
        ["Категория", "Технология", "Назначение"],
        [
            ("Фреймворк", "Vue 3", "SPA, мастер проектирования"),
            ("Язык", "TypeScript", "Типизация"),
            ("Сборка", "Vite 6", "Dev-сервер, production-сборка"),
            ("Маршрутизация", "Vue Router 4", "Навигация между экранами"),
            ("3D", "Three.js", "Просмотр кровли в 3D"),
            ("PDF", "pdfjs-dist", "Предпросмотр подложек PDF"),
            ("2D-редактор", "HTML5 Canvas / SVG", "Схема кровли, датчики, теплокарта"),
            ("Карты", "Яндекс.Карты JavaScript API", "Адрес на карте, геопривязка"),
            ("UI-прототип", "TN Design System", "Референс дизайна (React-прототип)"),
        ],
        [18, 28, 45],
    )

    ws3 = wb.create_sheet("Backend")
    write_table(
        ws3,
        1,
        ["Категория", "Технология", "Назначение"],
        [
            ("Runtime", "Go 1.22", "REST API"),
            ("HTTP", "chi v5", "Маршруты, middleware"),
            ("CORS", "go-chi/cors", "Доступ с фронтенда"),
            ("БД-драйвер", "pgx v5", "PostgreSQL"),
            ("S3", "minio-go v7", "Файловое хранилище"),
            ("PDF", "gofpdf", "Экспорт PDF-схем"),
            ("Excel", "excelize v2", "Экспорт Excel-отчётов"),
            ("ID", "google/uuid", "UUID проектов"),
            ("Пакет analytics", "Go", "Снеговые мешки, сетка нагрузки, климат, датчики"),
            ("Пакет export", "Go", "JSON / PDF / Excel"),
            ("Пакет store", "Go", "PostgreSQL"),
            ("Пакет storage", "Go", "MinIO/S3"),
            ("Пакет handler", "Go", "REST, геокодинг"),
        ],
        [18, 22, 45],
    )

    ws4 = wb.create_sheet("Инфраструктура")
    write_table(
        ws4,
        1,
        ["Компонент", "Технология", "Детали"],
        [
            ("СУБД", "PostgreSQL 16", "Основное хранилище"),
            ("ГИС", "PostGIS 3.4", "postgis/postgis:16-3.4"),
            ("Расширения", "pgcrypto", "UUID, криптография"),
            ("Данные", "JSONB", "Геометрия, климат, расчёты"),
            ("Кэш/очередь", "Redis 7", "Alpine, в Docker"),
            ("Файлы", "MinIO", "S3-совместимое хранилище"),
            ("Контейнеры", "Docker + Docker Compose", "Локальный и prod-like запуск"),
            ("Веб-сервер", "nginx 1.27", "Статика фронтенда"),
            ("Образы", "Alpine Linux", "Runtime API и web"),
        ],
        [20, 28, 42],
    )

    ws5 = wb.create_sheet("Интеграции")
    write_table(
        ws5,
        1,
        ["Сервис", "Использование"],
        [
            ("Яндекс.Карты / Геокодер", "Карта, адрес, координаты"),
            ("Nominatim (OpenStreetMap)", "Резервный геокодинг на backend"),
            ("ПИМ / QRC API (ТехноНИКОЛЬ)", "Системы кровли, λ, μ — теплотехника"),
        ],
        [32, 55],
    )

    ws6 = wb.create_sheet("Нормативы")
    write_table(
        ws6,
        1,
        ["Документ", "Применение"],
        [
            ("СП 20.13330.2016", "Снеговые нагрузки, снеговые мешки, пригрузы у парапетов"),
            ("СНиП / климатические справочники", "Снеговые и ветровые районы, роза ветров"),
            ("СП 230", "Теплотехнический расчёт"),
        ],
        [35, 55],
    )

    ws7 = wb.create_sheet("Инструменты")
    write_table(
        ws7,
        1,
        ["Инструмент", "Назначение"],
        [
            ("LikeC4", "C4-модель архитектуры (likec4/)"),
            ("Python 3", "Вспомогательные скрипты, синхронизация данных"),
        ],
        [28, 55],
    )

    ws8 = wb.create_sheet("Запланировано")
    write_table(
        ws8,
        1,
        ["Технология", "Назначение", "Статус"],
        [
            ("Worker (отдельный сервис)", "Автообводка DWG/DXF/PDF (CV)", "Архитектура / ТЗ"),
            ("go-geom, GEOS, GDAL", "Продвинутая геометрия", "Архитектура / ТЗ"),
            ("Корпоративный SSO", "Авторизация", "Архитектура / ТЗ"),
            ("Yandex Object Storage / AWS S3", "Production-альтернатива MinIO", "Архитектура / ТЗ"),
            ("Open-Meteo", "Роза ветров из архивных данных", "В документации, не в Go-MVP"),
        ],
        [32, 42, 22],
    )

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()
